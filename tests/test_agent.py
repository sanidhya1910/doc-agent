"""Routing, GPU accounting, and the tool layer.

The GPU-accounting tests are the quota guarantee from the design: they assert
that documents which should not touch the GPU genuinely do not.
"""

from __future__ import annotations

import tempfile

import pytest

from docagent import load, process
from docagent.router import plan_for, route
from docagent.state import DocKind, Document, Page, Table
from docagent.tools import TOOLS, call_tool, perceive, tool_schemas


def _gpu_steps(doc: Document) -> list[dict]:
    return [entry for entry in doc.trace if entry.get("gpu")]


# ---------------------------------------------------------------------------
# GPU accounting
# ---------------------------------------------------------------------------

def test_digital_pdf_costs_zero_gpu(fixture_path, workdir):
    doc = process(fixture_path("digital.pdf"), workdir=workdir)
    assert _gpu_steps(doc) == []
    assert doc.n_pages == 2
    assert doc.artifacts


@pytest.mark.parametrize("name", ["agreement.docx", "deck.pptx", "workbook.xlsx"])
def test_office_files_cost_zero_gpu(fixture_path, workdir, name):
    doc = process(fixture_path(name), workdir=workdir)
    assert _gpu_steps(doc) == []


@pytest.mark.parametrize("name", ["deck.pptx", "workbook.xlsx"])
def test_office_tables_reach_the_spreadsheet(fixture_path, workdir, name):
    """The end the user cares about: a table in, a populated workbook out."""
    import openpyxl

    doc = process(fixture_path(name), workdir=workdir)
    assert doc.all_tables()
    (xlsx,) = [a.path for a in doc.artifacts if a.kind == "xlsx"]
    workbook = openpyxl.load_workbook(xlsx)
    assert "Manifest" in workbook.sheetnames
    data_sheets = [s for s in workbook.sheetnames if s != "Manifest"]
    assert data_sheets
    rows = [[c.value for c in r] for r in workbook[data_sheets[0]].iter_rows()]
    assert len(rows) >= 2 and all(rows[0])


def test_trace_records_every_tool_call(fixture_path, workdir):
    doc = process(fixture_path("digital.pdf"), workdir=workdir)
    called = [entry["tool"] for entry in doc.trace]
    assert "classify_pages" in called
    assert "extract_tables" in called


# ---------------------------------------------------------------------------
# routing
# ---------------------------------------------------------------------------

def _doc_of_kind(kind: DocKind, *, tables: bool = False,
                 workdir: str | None = None) -> Document:
    page = Page(index=0, kind=kind, kind_confidence=0.9, has_text_layer=True,
                markdown="body text", backend="test")
    if tables:
        page.tables = [Table(header=["A"], rows=[["1"]], page_no=1)]
    # Never default the workdir to "." -- tools that write artifacts would
    # scatter them through the repo root.
    return Document(pages=[page], workdir=workdir or tempfile.mkdtemp())


def test_router_plans_never_write_files():
    """Export is the bundle's job -- writing inside the GPU session wastes quota."""
    for kind in (DocKind.TABLE, DocKind.INVOICE, DocKind.FORM, DocKind.PROSE):
        plan = plan_for(_doc_of_kind(kind)) or []
        assert not any(name.startswith("write_") for name, _ in plan)


@pytest.mark.parametrize(
    "kind,expected_tools",
    [
        (DocKind.TABLE, set()),          # perception already produced the tables
        (DocKind.INVOICE, {"extract_fields"}),
        (DocKind.FORM, {"extract_fields"}),
        (DocKind.PROSE, {"summarize"}),
    ],
)
def test_router_picks_the_expected_plan(kind, expected_tools):
    plan = plan_for(_doc_of_kind(kind))
    assert plan is not None
    assert expected_tools <= {name for name, _ in plan}


@pytest.mark.parametrize("kind", [DocKind.MIXED, DocKind.UNKNOWN])
def test_router_defers_on_ambiguous_documents(kind):
    """Ambiguity is the planner's job; the router must not guess.

    ``None`` means defer, which is distinct from an empty plan -- a plain table
    needs no further tools but is still handled by the router.
    """
    assert plan_for(_doc_of_kind(kind)) is None


def test_empty_plan_is_not_a_deferral():
    assert plan_for(_doc_of_kind(DocKind.TABLE)) == []


def test_router_defers_when_the_user_gave_an_instruction():
    doc = _doc_of_kind(DocKind.TABLE)
    doc.instruction = "just give me the plain text"
    assert route(doc) is None


def test_router_runs_without_an_instruction(workdir):
    doc = _doc_of_kind(DocKind.TABLE, tables=True, workdir=workdir)
    reply = route(doc)
    assert reply is not None
    assert "table" in reply.lower()


def test_router_path_uses_no_planner(fixture_path, workdir):
    """The unambiguous case must never reach the planner model."""
    doc = load(fixture_path("agreement.docx"), workdir=workdir)
    doc.workdir = workdir
    perceive(doc)
    assert route(doc) is not None
    assert not any(entry["tool"] == "plan" for entry in doc.trace)


def test_empty_document_defers():
    assert plan_for(Document()) is None


# ---------------------------------------------------------------------------
# the tool layer
# ---------------------------------------------------------------------------

def test_unknown_tool_returns_an_observation_not_an_exception(workdir):
    result = call_tool(Document(workdir=workdir), "no_such_tool", {})
    assert result.startswith("error: no such tool")


def test_bad_arguments_return_an_observation(workdir):
    doc = _doc_of_kind(DocKind.PROSE, workdir=workdir)
    result = call_tool(doc, "extract_tables", {"nonsense": 1})
    assert result.startswith("error: bad arguments")


def test_tool_failure_is_reported_not_raised(monkeypatch, workdir):
    doc = _doc_of_kind(DocKind.PROSE, workdir=workdir)

    def boom(*args, **kwargs):
        raise RuntimeError("kaboom")

    monkeypatch.setattr(TOOLS["extract_tables"], "fn", boom)
    assert "kaboom" in call_tool(doc, "extract_tables", {})


@pytest.mark.parametrize("spec,expected", [
    (None, [1, 2, 3]),
    ("all", [1, 2, 3]),
    (2, [2]),
    ([1, 3], [1, 3]),
    ("1, 3", [1, 3]),
])
def test_page_specs_are_tolerated(spec, expected):
    """The planner is not reliable about page-spec shape, so all forms work."""
    doc = Document(pages=[Page(index=i) for i in range(3)])
    assert [p.page_no for p in doc.resolve_pages(spec)] == expected


def test_tool_schemas_are_well_formed():
    schemas = tool_schemas()
    assert len(schemas) == len(TOOLS)
    for schema in schemas:
        function = schema["function"]
        assert function["name"] and function["description"]
        assert function["parameters"]["type"] == "object"


def test_write_xlsx_without_tables_explains_itself(workdir):
    doc = _doc_of_kind(DocKind.PROSE, workdir=workdir)
    assert "extract_tables" in call_tool(doc, "write_xlsx", {})


# ---------------------------------------------------------------------------
# document-level views
# ---------------------------------------------------------------------------

def test_dominant_kind_is_mixed_when_no_type_wins():
    doc = Document(pages=[
        Page(index=0, kind=DocKind.TABLE, kind_confidence=0.8),
        Page(index=1, kind=DocKind.PROSE, kind_confidence=0.8),
    ])
    assert doc.dominant_kind() is DocKind.MIXED


def test_dominant_kind_tolerates_one_outlier_page():
    doc = Document(pages=[
        Page(index=0, kind=DocKind.TABLE, kind_confidence=0.95),
        Page(index=1, kind=DocKind.TABLE, kind_confidence=0.95),
        Page(index=2, kind=DocKind.PROSE, kind_confidence=0.3),
    ])
    assert doc.dominant_kind() is DocKind.TABLE


@pytest.mark.parametrize("raw,expected", [
    ("invoice", DocKind.INVOICE),
    ("  Handwriting.\n", DocKind.HANDWRITING),
    ("id-document", DocKind.ID_DOCUMENT),
    ("this is a receipt", DocKind.RECEIPT),
    ("gibberish", DocKind.UNKNOWN),
])
def test_dockind_parsing_of_loose_model_output(raw, expected):
    assert DocKind.parse(raw) is expected


def test_brief_excludes_full_page_text():
    """The planner budget matters: it sees a summary, never the whole page."""
    page = Page(index=0, kind=DocKind.PROSE, kind_confidence=0.9,
                markdown="word " * 500, has_text_layer=True)
    brief = Document(pages=[page]).brief(preview_chars=100)
    assert len(brief) < 400
    assert "p1: prose" in brief


def test_process_with_no_readable_pages_warns_and_returns(workdir, tmp_path):
    bogus = tmp_path / "x.rtf"
    bogus.write_text("no", encoding="utf-8")
    doc = process(str(bogus), workdir=workdir)
    assert doc.n_pages == 0
    assert doc.warnings
